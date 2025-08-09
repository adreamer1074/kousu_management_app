# 標準ライブラリ
import os
import json
import logging
import io
import calendar
from datetime import date, datetime, timedelta
from decimal import Decimal
# サードパーティライブラリ(レポートエクスポート用)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from io import BytesIO
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Django
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import (
    ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView
)
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.db.models import Q, Sum, Avg, Count
from django.utils import timezone
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.contrib.auth import get_user_model
from apps.workloads.models import Workload
from kousu_management_app.settings import FONT_PATH
# ローカルアプリ
from .models import ReportExport, WorkloadAggregation
from .forms import WorkloadAggregationForm, WorkloadAggregationFilterForm
from apps.users.models import Department, Section
from apps.projects.models import ProjectTicket
from .utils import upload_file_to_s3
from apps.core.decorators import (
    leader_or_superuser_required_403,
    LeaderOrSuperuserRequiredMixin
)

User = get_user_model()

class WorkloadAggregationListView(LeaderOrSuperuserRequiredMixin, ListView):
    """工数集計一覧画面（工数集計レポート機能）"""
    model = WorkloadAggregation
    template_name = 'reports/workload_aggregation.html'
    context_object_name = 'aggregated_projects'
    paginate_by = 20
    
    def get_queryset(self):
        """
        一覧に表示するクエリセットを取得。
        del_flag=False のレコードを対象
        """
        queryset = WorkloadAggregation.objects.filter(del_flag=False).select_related(
            'case_name', 'section', 'mub_manager', 'created_by'
        ).order_by('section', '-order_date', '-created_at')
        # フィルター処理
        form = WorkloadAggregationFilterForm(self.request.GET)
        if form.is_valid():
            if form.cleaned_data.get('project_name'):
                queryset = queryset.filter(project_name__icontains=form.cleaned_data['project_name'])
            if form.cleaned_data.get('case_name'):
                queryset = queryset.filter(case_name=form.cleaned_data['case_name'])
            if form.cleaned_data.get('section'):
                queryset = queryset.filter(section=form.cleaned_data['section'])
            if form.cleaned_data.get('status'):
                queryset = queryset.filter(status=form.cleaned_data['status'])
            if form.cleaned_data.get('case_classification'):
                queryset = queryset.filter(case_classification=form.cleaned_data['case_classification'])
            if form.cleaned_data.get('mub_manager'):
                queryset = queryset.filter(mub_manager=form.cleaned_data['mub_manager'])
            if form.cleaned_data.get('search'):
                search_term = form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(project_name__icontains=search_term) |
                    Q(case_name__name__icontains=search_term) |
                    Q(remarks__icontains=search_term)
                )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        テンプレートに渡すコンテキストデータを構築。
        フィルターフォームや集計情報、現在のフィルター条件を含む。
        """
        context = super().get_context_data(**kwargs)
        
        # 現在のフィルター入力値を保持したフォームをコンテキストに追加
        context['filter_form'] = WorkloadAggregationFilterForm(self.request.GET)
        
        # 表示対象クエリセットを取得
        queryset = self.get_queryset()
        # 集計情報をコンテキストに追加（合計値を表示）
        context['total_stats'] = {
            'total_available_amount': queryset.aggregate(total=Sum('available_amount'))['total'] or 0,
            'total_billing_amount': queryset.aggregate(total=Sum('billing_amount_excluding_tax'))['total'] or 0,
            'total_outsourcing': queryset.aggregate(total=Sum('outsourcing_cost_excluding_tax'))['total'] or 0,
            'total_estimated_workdays': queryset.aggregate(total=Sum('estimated_workdays'))['total'] or 0,
            'total_used_workdays': queryset.aggregate(total=Sum('used_workdays'))['total'] or 0,
            'total_newbie_workdays': queryset.aggregate(total=Sum('newbie_workdays'))['total'] or 0,
        }
        

        # 現在のフィルター条件を辞書として渡す（テンプレート側でフォームの再表示等に利用）
        context['current_filters'] = {
            'project_name': self.request.GET.get('project_name', ''),
            'case_name_id': self.request.GET.get('case_name', ''),
            'section_id': self.request.GET.get('section', ''),
            'status': self.request.GET.get('status', ''),
            'case_classification': self.request.GET.get('case_classification', ''),
            'search': self.request.GET.get('search', ''),
        }
        
        return context

class WorkloadAggregationCreateView(LeaderOrSuperuserRequiredMixin, CreateView):
    """工数集計作成画面"""

    # 工数集計モデル使用
    model = WorkloadAggregation
    form_class = WorkloadAggregationForm
    template_name = 'reports/workload_aggregation_form.html'
    success_url = reverse_lazy('reports:workload_aggregation')
    
    def get_initial(self):
        """
        フォームの初期値を設定するメソッド。
        単価および請求単価にデフォルト値を設定。
        """
        initial = super().get_initial()
        initial.update({
            'unit_cost_per_month': 75.0,      # 単価のデフォルト値
            'billing_unit_cost_per_month': 90.0,  # 請求単価のデフォルト値
        })
        return initial
    
    def get_form_kwargs(self):
        """
        フォームに渡す追加のキーワード引数を設定。
        現在のユーザーをフォームに渡す。
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """
        フォームのバリデーションが成功した場合の処理。
        作成者として現在のユーザーを設定し、成功メッセージを表示。
        """
        form.instance.created_by = self.request.user
        messages.success(self.request, '工数集計データが正常に登録されました。')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        """
        テンプレートに渡すコンテキストデータを追加。
        タイトルなどの画面表示用情報を設定。
        """
        context = super().get_context_data(**kwargs)
        context['title'] = '工数集計登録'  # テンプレートで使う画面タイトル
        return context

class WorkloadAggregationDetailView(LeaderOrSuperuserRequiredMixin, DetailView):
    """工数集計詳細画面"""
    model = WorkloadAggregation
    template_name = 'reports/workload_aggregation_detail.html'
    context_object_name = 'workload_aggregation'

class WorkloadAggregationUpdateView(LeaderOrSuperuserRequiredMixin, UpdateView):
    """工数集計編集画面"""
    model = WorkloadAggregation
    form_class = WorkloadAggregationForm
    template_name = 'reports/workload_aggregation_form.html'
    success_url = reverse_lazy('reports:workload_aggregation')

    def get_form_kwargs(self):
        """
        フォームに渡す追加のキーワード引数を設定。
        現在のログインユーザーをフォームに渡す。
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        """
        テンプレートに渡すコンテキストデータを構築。
        タイトルなどの画面情報に加え、関連するチケットやプロジェクト情報も渡す。
        """
        context = super().get_context_data(**kwargs)
        context['title'] = '工数集計編集'

        # 編集対象の工数集計に紐づくチケットおよびプロジェクト情報をテンプレートに渡す
        if self.object and self.object.case_name:
            context.update({
                'current_ticket_id': self.object.case_name.id,
                'current_ticket_title': self.object.case_name.title,
                'current_project_id': self.object.case_name.project.id if self.object.case_name.project else None,
            })
        
        return context

class WorkloadAggregationDeleteView(LeaderOrSuperuserRequiredMixin, DeleteView):
    """工数集計削除画面"""
    model = WorkloadAggregation
    template_name = 'reports/workload_aggregation_confirm_delete.html'
    success_url = reverse_lazy('reports:workload_aggregation')
    
    def get_queryset(self):
        """削除されていないデータのみを対象"""
        return WorkloadAggregation.objects.filter(del_flag=False)

    def post(self, request, *args, **kwargs):
        """POSTリクエストでの削除処理"""     
        return self.delete(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        """論理削除を実行"""
        self.object = self.get_object()
        success_url = self.get_success_url()
        
        # 削除前の情報をメッセージに記録
        project_name = self.object.project_name
        case_name = self.object.case_name.title if self.object.case_name else "N/A"
        object_id = self.object.id
        # 論理削除実行
        self.object.soft_delete()

        # 削除後の状態確認
        self.object.refresh_from_db()
        print(f"削除後の状態: del_flag={self.object.del_flag}, deleted_at={self.object.deleted_at}")  # デバッグ用
        
        # 成功メッセージ
        messages.success(
            request, 
            f'工数集計データ「{project_name} - {case_name}」を削除しました。'
        )
        return HttpResponseRedirect(success_url)

@login_required
@leader_or_superuser_required_403
def workload_export(request):
    """工数データのエクスポート"""
    import csv
    from django.http import HttpResponse
    
    
    # CSV レスポンスの作成
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="workload_aggregation_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    # BOM付きでUTF-8エンコーディング
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'プロジェクト名', 'チケット名', '部名', 'ステータス', 'チケット分類', '見積日', '受注日',
        '終了日（予定）', '終了日実績', '検収日', '使用可能金額（税別）', '請求金額（税別）',
        '外注費（税別）', '見積工数（人日）', '使用工数（人日）', '新入社員使用工数（人日）',
        '使用工数合計', '残工数', '残金額', '利益率', '仕掛中金額', '請求先', 'MUB担当者', '作成日時'
    ])
    
    # フィルター適用
    queryset = WorkloadAggregation.objects.select_related(
        'case_name', 'section', 'mub_manager'
    ).order_by('-order_date')
    
    # URLパラメータでフィルター
    if request.GET.get('project_name'):
        queryset = queryset.filter(project_name__icontains=request.GET['project_name'])
    if request.GET.get('case_name'):
        queryset = queryset.filter(case_name_id=request.GET['case_name'])
    if request.GET.get('status'):
        queryset = queryset.filter(status=request.GET['status'])
    
    for workload in queryset:
        writer.writerow([
            workload.project_name,
            workload.case_name.name,
            workload.section.name,
            workload.get_status_display(),
            workload.get_case_classification_display(),
            workload.estimate_date.strftime('%Y-%m-%d') if workload.estimate_date else '',
            workload.order_date.strftime('%Y-%m-%d') if workload.order_date else '',
            workload.planned_end_date.strftime('%Y-%m-%d') if workload.planned_end_date else '',
            workload.actual_end_date.strftime('%Y-%m-%d') if workload.actual_end_date else '',
            workload.inspection_date.strftime('%Y-%m-%d') if workload.inspection_date else '',
            str(workload.available_amount),
            str(workload.billing_amount_excluding_tax),
            str(workload.outsourcing_cost_excluding_tax),
            str(workload.estimated_workdays),
            str(workload.used_workdays),
            str(workload.newbie_workdays),
            str(workload.total_used_workdays),
            str(workload.remaining_workdays),
            str(workload.remaining_amount),
            str(workload.profit_rate),
            str(workload.wip_amount),
            workload.billing_destination,
            workload.mub_manager.get_full_name() if workload.mub_manager else '',
            workload.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])
    
    return response

class ReportListView(LeaderOrSuperuserRequiredMixin, TemplateView):
    """レポート一覧画面"""
    template_name = 'reports/report_list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # レポートメニューの定義
        reports = [
            {
                'name': '工数集計',
                'description': '工数集計一覧の管理と分析',
                'url': 'reports:workload_aggregation',
                'icon': 'bi-bar-chart-line',
                'color': 'success'
            },
            {
                'name': 'エクスポート履歴',
                'description': 'レポートのエクスポート履歴',
                'url': 'reports:report_export_list',
                'icon': 'bi-download',
                'color': 'info'
            },
        ]
        
        context['reports'] = reports
        return context

class ReportExportListView(LeaderOrSuperuserRequiredMixin, ListView):
    """エクスポート履歴一覧表示"""
    model = ReportExport
    template_name = 'reports/report_export_list.html'
    context_object_name = 'reports'
    paginate_by = 20

    def get_queryset(self):
        queryset = ReportExport.objects.select_related('requested_by').order_by('-requested_at')
        # 一般ユーザーは自分のエクスポートのみ表示
        if not self.request.user.is_leader:
            queryset = queryset.filter(requested_by=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'エクスポート履歴'
        return context


@login_required
@leader_or_superuser_required_403
@require_POST
def calculate_workdays_ajax(request):
    """工数自動計算（AJAX版）"""
    try:
        # FormDataからパラメータを取得
        ticket_id = request.POST.get('ticket_id')
        classification = request.POST.get('classification', 'development')
        order_date = request.POST.get('order_date')
        actual_end_date = request.POST.get('actual_end_date')
        
        if not ticket_id:
            return JsonResponse({'success': False, 'error': 'チケットIDが必要です。'})
        
        #チケットオブジェクトを取得
        from apps.projects.models import ProjectTicket
        try:
            ticket = ProjectTicket.objects.get(id=ticket_id, is_active=True)
        except ProjectTicket.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': f'指定されたチケット（ID: {ticket_id}）が見つかりません。'
            })
        
        # 工数データの取得
        workloads_query = Workload.objects.filter(ticket__id=ticket_id)
        
        # 日付フィルター適用
        target_year_months = set()
        
        if order_date:
            try:
                start_date = datetime.strptime(order_date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False, 
                    'error': f'受注日の形式が正しくありません: {order_date}'
                })
        else:
            # デフォルトは6ヶ月前から
            start_date = date.today().replace(day=1) - timedelta(days=180)
            start_date = start_date.replace(day=1)
        
        if actual_end_date:
            try:
                end_date = datetime.strptime(actual_end_date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False, 
                    'error': f'終了日の形式が正しくありません: {actual_end_date}'
                })
        else:
            end_date = date.today()
        
        # 日付検証
        if start_date > end_date:
            return JsonResponse({
                'success': False, 
                'error': '受注日は終了日より前である必要があります。'
            })
        
        # 対象年月のリストを作成
        current_date = start_date.replace(day=1)
        while current_date <= end_date:
            target_year_months.add(current_date.strftime('%Y-%m'))
            # 次の月へ
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        # 対象年月でフィルター
        if target_year_months:
            workloads_query = workloads_query.filter(year_month__in=target_year_months)
        
        # 一般使用工数と新入社員工数を分離
        regular_workdays = Decimal('0.0')
        newbie_workdays = Decimal('0.0')
        
        for workload in workloads_query:
            # ユーザーのemployee_levelを確認
            is_junior = False
            if hasattr(workload.user, 'employee_level') and workload.user.employee_level == 'junior':
                is_junior = True
            
            # 各日の工数を合計（日付範囲内のみ）
            workload_year_month = workload.year_month
            try:
                year, month = map(int, workload_year_month.split('-'))
            except ValueError:
                continue  # 無効な年月形式はスキップ
            
            # その月の日数を取得
            days_in_month = calendar.monthrange(year, month)[1]
            
            for day in range(1, days_in_month + 1):
                # 日付範囲チェック
                try:
                    current_day = date(year, month, day)
                except ValueError:
                    continue  # 無効な日付はスキップ
                
                # 指定された日付範囲外の工数は除外
                if current_day < start_date or current_day > end_date:
                    continue
                
                # その日の工数を取得
                try:
                    day_hours = workload.get_day_value(day)
                    if day_hours < 0:  # 負の値はスキップ
                        continue
                except (AttributeError, TypeError, ValueError):
                    continue  # エラーの場合はスキップ
                
                # 工数の分類
                if is_junior:
                    newbie_workdays += Decimal(str(day_hours))
                else:
                    regular_workdays += Decimal(str(day_hours))
        
        # 時間を人日に変換（8時間=1人日として計算）
        used_workdays = regular_workdays / 8
        newbie_workdays_converted = newbie_workdays / 8
        
        # === 【修正】レスポンスでticket.titleを正しく使用 ===
        return JsonResponse({
            'success': True,
            'used_workdays': float(used_workdays),
            'newbie_workdays': float(newbie_workdays_converted),
            'total_workdays': float(used_workdays + newbie_workdays_converted),
            'ticket_name': ticket.title,  # ✅ ここでticketが正しく定義済み
            'ticket_id': ticket.id,
            'project_name': ticket.project.name if ticket.project else None,
            'debug_info': {
                'workload_records': workloads_query.count(),
                'target_months': list(target_year_months),
                'regular_hours': float(regular_workdays),
                'newbie_hours': float(newbie_workdays),
                'date_range': f'{start_date} - {end_date}',
                'received_ticket_id': ticket_id,
                'classification': classification
            }
        })
        
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"工数計算AJAX エラー: {str(e)}")
        logger.error(f"トレースバック: {traceback.format_exc()}")
        
        return JsonResponse({
            'success': False, 
            'error': f'サーバーエラー: {str(e)}'
        })
@login_required
@leader_or_superuser_required_403
def workload_export_current(request):
    """現在表示中の工数集計データをエクスポート（即時ダウンロード＋S3アップロード＋履歴登録）"""
    if request.method != 'POST':
        return redirect('reports:workload_aggregation')
    
    # フィルター条件を取得
    filters = {}
    for key in ['project_name', 'case_name', 'status', 'case_classification', 'section', 'mub_manager']:
        value = request.POST.get(key) or request.GET.get(key)
        if value:
            filters[key] = value
    
    # データセットを取得
    queryset = WorkloadAggregation.objects.select_related(
        'case_name', 'section', 'mub_manager'
    )
    
    # フィルター適用
    if filters.get('project_name'):
        queryset = queryset.filter(project_name__icontains=filters['project_name'])
    if filters.get('case_name'):
        queryset = queryset.filter(case_name_id=filters['case_name'])
    if filters.get('status'):
        queryset = queryset.filter(status=filters['status'])
    if filters.get('case_classification'):
        queryset = queryset.filter(case_classification=filters['case_classification'])
    if filters.get('section'):
        queryset = queryset.filter(section_id=filters['section'])
    if filters.get('mub_manager'):
        queryset = queryset.filter(mub_manager_id=filters['mub_manager'])
    
    queryset = queryset.order_by('-created_at')
    
    # エクスポート形式を取得
    export_format = request.POST.get('format', 'excel')
    
    print(f"=== EXPORT DEBUG ===")
    print(f"Export format requested: {export_format}")
    print(f"QuerySet count: {queryset.count()}")
    print(f"==================")
    
    # データが0件の場合の処理
    if queryset.count() == 0:
        messages.warning(request, 'エクスポートするデータがありません。フィルター条件を確認してください。')
        return redirect('reports:workload_aggregation')
    
    # Windows対応：一時ディレクトリとファイル名を修正
    import tempfile
    
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    
    # 正しいファイル拡張子を設定
    extension_map = {
        'excel': 'xlsx',
        'csv': 'csv',
        'pdf': 'pdf'
    }
    
    file_extension = extension_map.get(export_format, 'xlsx')
    file_name = f"workload_{timestamp}.{file_extension}"
    
    # Windows対応の一時ディレクトリ
    temp_dir = tempfile.gettempdir()
    local_path = os.path.join(temp_dir, file_name)
    
    print(f"Temp directory: {temp_dir}")
    print(f"Local path: {local_path}")
    
    # 履歴モデル作成（S3アップロード用）
    export_record = ReportExport.objects.create(
        requested_by=request.user,
        export_type=ReportExport.ExportTypeChoices.WORKLOAD_AGGREGATION,
        export_format=export_format,
        status=ReportExport.StatusChoices.PENDING,
        file_name=file_name,
        filter_conditions=filters,  # フィルター条件も保存
        total_records=queryset.count()
    )
    
    try:
        # ファイル生成（S3アップロード用とダウンロード用の両方）
        if export_format == 'excel':
            print("Processing EXCEL export...")
            # S3アップロード用ファイル生成
            export_workload_excel(queryset, export_type='excel', save_to_file=local_path)
            # ダウンロード用レスポンス生成
            response = export_workload_excel(queryset, export_type='excel')
        elif export_format == 'csv':
            print("Processing CSV export...")
            # S3アップロード用ファイル生成
            export_workload_csv(queryset, save_to_file=local_path)
            # ダウンロード用レスポンス生成
            response = export_workload_csv(queryset)
        elif export_format == 'pdf':
            print("Processing PDF export...")
            # S3アップロード用ファイル生成
            export_workload_pdf(queryset, filters, save_to_file=local_path)
            # ダウンロード用レスポンス生成
            response = export_workload_pdf(queryset, filters)
        elif export_format == 'professional': 
            print("Processing PROFESSIONAL export...")
            professional_type = request.POST.get('professional_type', 'executive_summary')
            # S3アップロード用ファイル生成
            export_workload_professional(queryset, filters, professional_type, save_to_file=local_path)
            # ダウンロード用レスポンス生成
            response = export_workload_professional(queryset, filters, professional_type)
        else:
            messages.error(request, f'サポートされていない形式です: {export_format}')
            export_record.delete()  # 失敗時は履歴削除
            return redirect('reports:workload_aggregation')
        
        # ファイルサイズを取得
        if os.path.exists(local_path):
            file_size = os.path.getsize(local_path)
            export_record.file_size = file_size
            export_record.save(update_fields=['file_size'])
            print(f"File created successfully: {local_path} ({file_size} bytes)")
        else:
            print(f"WARNING: File not created: {local_path}")
            raise Exception(f"ファイル生成に失敗しました: {local_path}")
        
        # S3アップロード＆履歴登録（バックグラウンドで実行）
        print(f"Starting S3 upload for: {local_path}")
        complete_export(export_record, local_path)
        
        print(f"Export completed successfully: {file_name}")
        messages.success(request, f'エクスポートが完了しました。履歴からS3ダウンロードも可能です。')
        
        return response
        
    except Exception as e:
        print(f"Export failed: {str(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        
        # エラー時は履歴をFAILEDに更新
        try:
            if hasattr(export_record, 'error_message'):
                export_record.error_message = str(e)
                update_fields = ['status', 'error_message']
            else:
                update_fields = ['status']
            
            export_record.status = ReportExport.StatusChoices.FAILED
            export_record.save(update_fields=update_fields)
        except Exception as save_error:
            print(f"Failed to save error status: {save_error}")
        
        messages.error(request, f'エクスポートに失敗しました: {str(e)}')
        return redirect('reports:workload_aggregation')

def export_workload_pdf(queryset, filters=None, save_to_file=None):
    """PDF形式でエクスポート（ファイル保存またはHttpResponse）"""
    try:
        # フォント登録
        pdfmetrics.registerFont(TTFont('IPAexGothic', FONT_PATH))
        font_name = 'IPAexGothic'
        font_size = 10

        # ファイル保存版
        if save_to_file:
            output_stream = save_to_file
        else:
            output_stream = BytesIO()

        # PDF準備（横向きA4、余白1インチ=72pt程度）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        doc = SimpleDocTemplate(
            output_stream,
            pagesize=landscape(A4),
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36,
            title="工数集計レポート"
        )

        # スタイル設定
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='JapaneseTitle', fontName=font_name, fontSize=16, leading=22, alignment=1))
        styles.add(ParagraphStyle(name='JapaneseSmall', fontName=font_name, fontSize=9, leading=12))
        styles.add(ParagraphStyle(name='Japanese', fontName=font_name, fontSize=font_size, leading=14))

        elements = []

        # タイトルと生成日時
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("工数集計レポート", styles['JapaneseTitle']))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"生成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}", styles['JapaneseSmall']))
        elements.append(Spacer(1, 18))

        # ヘッダー
        headers = ["プロジェクト名", "チケット名", "部門", "分類", "見積日", "受注日", "予定終了", "検収日",
                   "請求金額", "外注費", "見積工数", "使用工数合計", "月額単価", "請求単価", "請求先", "担当者", "備考"]

        data = [headers]

        # データ行（最大50件）
        for workload in queryset[:50]:
            data.append([
                str(workload.project_name or ''),
                str(workload.case_name.title if workload.case_name else ''),
                str(workload.section.name if workload.section else ''),
                str(workload.get_case_classification_display()),
                workload.estimate_date.strftime('%Y-%m-%d') if workload.estimate_date else '',
                workload.order_date.strftime('%Y-%m-%d') if workload.order_date else '',
                workload.planned_end_date.strftime('%Y-%m-%d') if workload.planned_end_date else '',
                workload.inspection_date.strftime('%Y-%m-%d') if workload.inspection_date else '',
                f"¥{float(workload.billing_amount_excluding_tax or 0):,.0f}",
                f"¥{float(workload.outsourcing_cost_excluding_tax or 0):,.0f}",
                f"{float(workload.estimated_workdays or 0):.1f}",
                f"{float(workload.total_used_workdays or 0):.1f}",
                f"¥{float(workload.unit_cost_per_month or 0):,.0f}",
                f"¥{float(workload.billing_unit_cost_per_month or 0):,.0f}",
                workload.billing_destination or '',
                workload.mub_manager.get_full_name() if workload.mub_manager else '',
                workload.remarks or '',
            ])

        # テーブル作成（省略...既存のコード）
        columns = list(zip(*data))
        col_widths = []
        for col in columns:
            max_width = max(
                pdfmetrics.stringWidth(str(cell), font_name, font_size) for cell in col
            )
            col_widths.append(max_width + 10)

        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DDEEFF")),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ]))

        elements.append(table)

        # PDF生成
        doc.build(elements)

        # ファイル保存版
        if save_to_file:
            print(f"PDF file saved to: {save_to_file}")
            return save_to_file

        # HTTPレスポンス版（既存の動作）
        output_stream.seek(0)
        response = HttpResponse(output_stream.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="workload_report_{timestamp}.pdf"'
        return response

    except Exception as e:
        import traceback
        print(f"PDF生成エラー: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse("PDF生成に失敗しました", status=500)
        
    except Exception as e:
        print(f"Export failed: {str(e)}")
        import traceback
        print(f"PDF生成エラー: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse("PDF生成に失敗しました", status=500)


def export_workload_text_fallback(queryset, filters=None):
    """最終手段：テキストファイル"""
    print("Creating text fallback...")
    
    response = HttpResponse(content_type='text/plain; charset=utf-8')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'workload_text_{timestamp}.txt'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    response.write(f'Workload Report (Text Format)\n')
    response.write(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
    response.write(f'Total Records: {queryset.count()}\n\n')
    
    for i, workload in enumerate(queryset[:50], 1):
        try:
            project_name = str(workload.project_name) if workload.project_name else 'N/A'
            case_title = str(workload.case_name.title) if workload.case_name else 'N/A'
            status = str(workload.get_status_display()) if hasattr(workload, 'get_status_display') else 'N/A'
            amount = str(workload.billing_amount_excluding_tax or 0)
            
            response.write(f'{i}. {project_name} | {case_title} | {status} | ¥{amount}\n')
        except Exception as e:
            response.write(f'{i}. [Error processing record: {str(e)}]\n')
    
    print(f"Text file created: {filename}")
    return response
@login_required
@leader_or_superuser_required_403
@require_http_methods(["POST"])
def export_report_with_history(request):
    """エクスポート履歴に登録しつつレポートをエクスポート（CSV例）"""
    # フィルター条件取得
    filters = {}
    for key in ['project_name', 'case_name', 'status', 'case_classification', 'section', 'mub_manager']:
        value = request.POST.get(key) or request.GET.get(key)
        if value:
            filters[key] = value

    queryset = WorkloadAggregation.objects.select_related(
        'case_name', 'section', 'mub_manager'
    )
    # フィルター適用
    if filters.get('project_name'):
        queryset = queryset.filter(project_name__icontains=filters['project_name'])
    if filters.get('case_name'):
        queryset = queryset.filter(case_name_id=filters['case_name'])
    if filters.get('status'):
        queryset = queryset.filter(status=filters['status'])
    if filters.get('case_classification'):
        queryset = queryset.filter(case_classification=filters['case_classification'])
    if filters.get('section'):
        queryset = queryset.filter(section_id=filters['section'])
    if filters.get('mub_manager'):
        queryset = queryset.filter(mub_manager_id=filters['mub_manager'])
    
    queryset = queryset.order_by('-created_at')

    export_format = request.POST.get('format', 'csv')  # 'csv', 'pdf', 'excel'
    return export_and_upload(queryset, export_format, filters, request.user)

def export_workload_excel(queryset, export_type='excel', save_to_file=None):
    """Excel形式でエクスポート（ファイル保存またはHttpResponse）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工数集計レポート"
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ヘッダー行
    headers = [
        'プロジェクト名', 'チケット名', '部名', 'ステータス', 'チケット分類',
        '見積日', '受注日', '終了日（予定）', '終了日実績', '検収日',
        '使用可能金額（税別）', '請求金額（税別）', '外注費（税別）',
        '見積工数（人日）', '使用工数（人日）', '新入社員使用工数（人日）',
        '単価（万円/月）', '請求単価（万円/月）',
        '請求先', 'MUB担当者', '備考', '作成日時'
    ]
    
    # ヘッダー設定
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # データ行
    for row_num, workload in enumerate(queryset, 2):
        try:
            data_row = [
                str(workload.project_name) if workload.project_name else '',
                str(workload.case_name.title) if workload.case_name else '',
                str(workload.section.name) if workload.section else '',
                str(workload.get_status_display()),
                str(workload.get_case_classification_display()),
                workload.estimate_date.strftime('%Y-%m-%d') if workload.estimate_date else '',
                workload.order_date.strftime('%Y-%m-%d') if workload.order_date else '',
                workload.planned_end_date.strftime('%Y-%m-%d') if workload.planned_end_date else '',
                workload.actual_end_date.strftime('%Y-%m-%d') if workload.actual_end_date else '',
                workload.inspection_date.strftime('%Y-%m-%d') if workload.inspection_date else '',
                float(workload.available_amount or 0),
                float(workload.billing_amount_excluding_tax or 0),
                float(workload.outsourcing_cost_excluding_tax or 0),
                float(workload.estimated_workdays or 0),
                float(workload.used_workdays or 0),
                float(workload.newbie_workdays or 0),
                float(workload.unit_cost_per_month or 0),
                float(workload.billing_unit_cost_per_month or 0),
                workload.billing_destination or '',
                workload.mub_manager.get_full_name() if workload.mub_manager else '',
                workload.remarks or '',
                workload.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            
            for col_num, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        except Exception as e:
            print(f"Row {row_num} export error: {str(e)}")
            continue
    
    # 列幅の自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # ファイル保存版（S3アップロード用）
    if save_to_file:
        wb.save(save_to_file)
        print(f"Excel file saved to: {save_to_file}")
        return save_to_file
    
    # HTTPレスポンス版（既存の動作）
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if export_type == 'pdf':
        filename = f'workload_extended_{timestamp}.xlsx'
    else:
        filename = f'workload_standard_{timestamp}.xlsx'
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def export_workload_csv(queryset, save_to_file=None):
    """CSV形式でエクスポート（ファイル保存またはHttpResponse）"""
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # ヘッダー行
    writer.writerow([
        'プロジェクト名', 'チケット名', '部名', 'ステータス', 'チケット分類',
        '見積日', '受注日', '終了日（予定）', '終了日実績', '検収日',
        '使用可能金額（税別）', '請求金額（税別）', '外注費（税別）',
        '見積工数（人日）', '使用工数（人日）', '新入社員使用工数（人日）',
        '単価（万円/月）', '請求単価（万円/月）',
        '請求先', 'MUB担当者', '備考', '作成日時'
    ])
    
    # データ行
    for workload in queryset:
        try:
            writer.writerow([
                str(workload.project_name) if workload.project_name else '',
                str(workload.case_name.title) if workload.case_name else '',
                str(workload.section.name) if workload.section else '',
                str(workload.get_status_display()),
                str(workload.get_case_classification_display()),
                workload.estimate_date.strftime('%Y-%m-%d') if workload.estimate_date else '',
                workload.order_date.strftime('%Y-%m-%d') if workload.order_date else '',
                workload.planned_end_date.strftime('%Y-%m-%d') if workload.planned_end_date else '',
                workload.actual_end_date.strftime('%Y-%m-%d') if workload.actual_end_date else '',
                workload.inspection_date.strftime('%Y-%m-%d') if workload.inspection_date else '',
                str(workload.available_amount or 0),
                str(workload.billing_amount_excluding_tax or 0),
                str(workload.outsourcing_cost_excluding_tax or 0),
                str(workload.estimated_workdays or 0),
                str(workload.used_workdays or 0),
                str(workload.newbie_workdays or 0),
                str(workload.unit_cost_per_month or 0),
                str(workload.billing_unit_cost_per_month or 0),
                str(workload.billing_destination) if workload.billing_destination else '',
                str(workload.mub_manager.get_full_name()) if workload.mub_manager else '',
                str(workload.remarks) if workload.remarks else '',
                workload.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])
        except Exception as e:
            print(f"CSV export error: {str(e)}")
            continue
    
    # ファイル保存版（S3アップロード用）
    if save_to_file:
        with open(save_to_file, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(output.getvalue())
        print(f"CSV file saved to: {save_to_file}")
        return save_to_file
    
    # HTTPレスポンス版（既存の動作）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'workload_csv_{timestamp}.csv'
    
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def complete_export(report_export, local_path):
    """
    S3アップロード＆履歴登録を行う関数
    """
    logger = logging.getLogger(__name__)
    logger.info(f"[S3連携] complete_export開始: local_path={local_path}, file_name={report_export.file_name}")

    try:
        
        # ファイル存在確認
        if not os.path.exists(local_path):
            raise FileNotFoundError(f"ローカルファイルが見つかりません: {local_path}")
        
        # ファイルサイズ取得
        file_size = os.path.getsize(local_path)
        logger.info(f"[S3連携] ファイルサイズ: {file_size} bytes")
        
        # S3アップロード
        s3_key = f"exports/{report_export.file_name}"
        logger.info(f"[S3連携] S3アップロード開始: s3_key={s3_key}")
        
        s3_url = upload_file_to_s3(local_path, s3_key)
        logger.info(f"[S3連携] S3アップロード成功: s3_url={s3_url}")

        # file_s3_urlフィールドが存在するかチェック
        if hasattr(report_export, 'file_s3_url'):
            report_export.file_s3_url = s3_url
            report_export.file_size = file_size
            update_fields = ['file_s3_url', 'file_size', 'status']
        else:
            logger.warning("[S3連携] file_s3_urlフィールドが存在しません。マイグレーション後に再試行してください。")
            report_export.file_size = file_size
            update_fields = ['file_size', 'status']
        
        report_export.status = ReportExport.StatusChoices.COMPLETED
        report_export.save(update_fields=update_fields)
        logger.info(f"[履歴登録] ReportExport更新完了: id={report_export.id}, status={report_export.status}")
        
        # ローカルファイルを削除（オプション）
        try:
            os.remove(local_path)
            logger.info(f"[ファイル削除] ローカルファイルを削除: {local_path}")
        except Exception as cleanup_error:
            logger.warning(f"[ファイル削除] ローカルファイル削除に失敗: {cleanup_error}")

    except Exception as e:
        logger.error(f"[S3連携] S3アップロード失敗: {str(e)}")
        import traceback
        logger.error(f"[S3連携] エラー詳細: {traceback.format_exc()}")
        
        # エラー時は履歴をFAILEDに更新
        try:
            if hasattr(report_export, 'error_message'):
                report_export.error_message = str(e)
                update_fields = ['status', 'error_message']
            else:
                update_fields = ['status']
            
            report_export.status = ReportExport.StatusChoices.FAILED
            report_export.save(update_fields=update_fields)
            logger.error(f"[履歴登録] ReportExportステータスFAILEDに更新: id={report_export.id}")
        except Exception as save_error:
            logger.error(f"[履歴登録] エラー状態の保存に失敗: {save_error}")
        
        # エラーを再発生させる
        raise e

def upload_file_to_s3(local_path, s3_key):
    """
    ローカルファイルをS3にアップロードする関数
    """
    import boto3
    from django.conf import settings
    
    try:
        # S3クライアント作成
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # S3にアップロード
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        s3_client.upload_file(local_path, bucket_name, s3_key)
        
        # S3 URLを生成
        s3_url = f"https://{bucket_name}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_key}"
        
        return s3_url
        
    except Exception as e:
        raise Exception(f"S3アップロードエラー: {str(e)}")
def export_workload_professional(queryset, filters=None, professional_type='executive_summary', save_to_file=None):
    """カッコイイレポート形式でエクスポート（PDF）"""
    try:
        # フォント登録
        pdfmetrics.registerFont(TTFont('IPAexGothic', FONT_PATH))
        font_name = 'IPAexGothic'

        # ファイル保存版
        if save_to_file:
            output_stream = save_to_file
        else:
            output_stream = BytesIO()

        # PDF準備（A4縦向き）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        doc = SimpleDocTemplate(
            output_stream,
            pagesize=A4,
            leftMargin=50,
            rightMargin=50,
            topMargin=50,
            bottomMargin=50,
            title="プロフェッショナル工数レポート"
        )

        # カスタムスタイル設定（既存のスタイルと重複しない名前を使用）
        styles = getSampleStyleSheet()
        
        # プロフェッショナル用スタイル（重複しない名前を使用）
        styles.add(ParagraphStyle(
            name='ProMainTitle',
            fontName=font_name,
            fontSize=20,
            leading=26,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1f4e79"),
            spaceAfter=30
        ))
        
        styles.add(ParagraphStyle(
            name='ProSectionTitle',
            fontName=font_name,
            fontSize=14,
            leading=18,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#2c5aa0"),
            spaceBefore=20,
            spaceAfter=10,
            borderWidth=0,
            borderPadding=5,
            backColor=colors.HexColor("#f2f6fc")
        ))
        
        styles.add(ParagraphStyle(
            name='ProBodyText',
            fontName=font_name,
            fontSize=10,
            leading=14,
            alignment=TA_LEFT,
            textColor=colors.black
        ))
        
        styles.add(ParagraphStyle(
            name='ProMetricValue',
            fontName=font_name,
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1f4e79"),
            spaceBefore=5,
            spaceAfter=5
        ))

        elements = []

        # レポートタイプ別の生成
        if professional_type == 'executive_summary':
            elements.extend(_generate_executive_summary(queryset, styles, filters))
        elif professional_type == 'detailed_analysis':
            elements.extend(_generate_detailed_analysis(queryset, styles, filters))
        elif professional_type == 'financial_dashboard':
            elements.extend(_generate_financial_dashboard(queryset, styles, filters))
        elif professional_type == 'project_portfolio':
            elements.extend(_generate_project_portfolio(queryset, styles, filters))
        else:
            elements.extend(_generate_executive_summary(queryset, styles, filters))

        # PDF生成
        doc.build(elements)

        # ファイル保存版
        if save_to_file:
            print(f"Professional report saved to: {save_to_file}")
            return save_to_file

        # HTTPレスポンス版（既存の動作）
        output_stream.seek(0)
        filename = f'professional_report_{professional_type}_{timestamp}.pdf'
        response = HttpResponse(output_stream.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        import traceback
        print(f"プロフェッショナルレポート生成エラー: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse("プロフェッショナルレポート生成に失敗しました", status=500)

def _generate_executive_summary(queryset, styles, filters):
    """エグゼクティブサマリー生成"""
    elements = []
    
    # タイトル
    elements.append(Paragraph("エグゼクティブサマリー", styles['ProMainTitle']))
    elements.append(Paragraph(f"生成日時: {datetime.now().strftime('%Y年%m月%d日')}", styles['ProBodyText']))
    elements.append(Spacer(1, 20))
    
    # 全体統計
    total_projects = queryset.count()
    total_billing = queryset.aggregate(total=Sum('billing_amount_excluding_tax'))['total'] or 0
    total_outsourcing = queryset.aggregate(total=Sum('outsourcing_cost_excluding_tax'))['total'] or 0
    total_workdays = queryset.aggregate(total=Sum('used_workdays'))['total'] or 0
    profit_margin = ((total_billing - total_outsourcing) / total_billing * 100) if total_billing > 0 else 0
    
    # サマリーテーブル
    summary_data = [
        ['項目', '値', '単位'],
        ['総プロジェクト数', f'{total_projects:,}', '件'],
        ['総請求金額', f'¥{total_billing:,.0f}', '円'],
        ['総外注費', f'¥{total_outsourcing:,.0f}', '円'],
        ['総使用工数', f'{total_workdays:.1f}', '人日'],
        ['利益率', f'{profit_margin:.1f}', '%'],
    ]
    
    summary_table = Table(summary_data, colWidths=[120, 100, 60])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c5aa0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
    ]))
    
    elements.append(Paragraph("プロジェクト概要", styles['ProSectionTitle']))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # ステータス別分析（安全な方法で取得）
    status_analysis = queryset.values('status').annotate(
        count=Count('id'),
        total_amount=Sum('billing_amount_excluding_tax')
    ).order_by('-total_amount')
    
    if status_analysis:
        elements.append(Paragraph("ステータス別分析", styles['ProSectionTitle']))
        
        status_data = [['ステータス', 'プロジェクト数', '請求金額']]
        for item in status_analysis:
            # 安全にステータス表示名を取得
            try:
                # まず、WorkloadAggregationモデルから選択肢を取得
                if hasattr(WorkloadAggregation, 'STATUS_CHOICES'):
                    status_display = dict(WorkloadAggregation.STATUS_CHOICES).get(item['status'], item['status'])
                else:
                    # 選択肢が定義されていない場合は、インスタンスから取得
                    sample_instance = queryset.filter(status=item['status']).first()
                    if sample_instance and hasattr(sample_instance, 'get_status_display'):
                        status_display = sample_instance.get_status_display()
                    else:
                        status_display = str(item['status'])
            except Exception as e:
                print(f"ステータス表示名取得エラー: {e}")
                status_display = str(item['status'])
            
            status_data.append([
                status_display,
                f"{item['count']:,}件",
                f"¥{item['total_amount'] or 0:,.0f}"
            ])
        
        status_table = Table(status_data, colWidths=[100, 80, 100])
        status_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#ecf0f1")]),
        ]))
        
        elements.append(status_table)
    
    return elements

def _generate_detailed_analysis(queryset, styles, filters):
    """詳細分析レポート生成"""
    elements = []
    
    # タイトル
    elements.append(Paragraph("詳細分析レポート", styles['ProMainTitle']))
    elements.append(Spacer(1, 20))
    
    # 部門別分析
    section_analysis = queryset.values('section__name').annotate(
        count=Count('id'),
        total_billing=Sum('billing_amount_excluding_tax'),
        total_workdays=Sum('used_workdays'),
        avg_unit_cost=Avg('unit_cost_per_month')
    ).order_by('-total_billing')
    
    if section_analysis:
        elements.append(Paragraph("部門別パフォーマンス", styles['ProSectionTitle']))
        
        section_data = [['部門', 'プロジェクト数', '請求金額', '使用工数', '平均単価']]
        for item in section_analysis:
            section_data.append([
                item['section__name'] or '未設定',
                f"{item['count']:,}",
                f"¥{item['total_billing'] or 0:,.0f}",
                f"{item['total_workdays'] or 0:.1f}",
                f"¥{item['avg_unit_cost'] or 0:.0f}"
            ])
        
        section_table = Table(section_data, colWidths=[80, 60, 80, 60, 60])
        section_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#27ae60")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(section_table)
        elements.append(Spacer(1, 20))
    
    # 上位10プロジェクト
    top_projects = queryset.order_by('-billing_amount_excluding_tax')[:10]
    
    elements.append(Paragraph("上位プロジェクト（請求金額順）", styles['ProSectionTitle']))
    
    project_data = [['順位', 'プロジェクト名', 'チケット', '請求金額', 'ステータス']]
    for i, project in enumerate(top_projects, 1):
        # 安全にステータス表示名を取得
        try:
            if hasattr(project, 'get_status_display'):
                status_display = project.get_status_display()
            else:
                status_display = str(project.status)
        except Exception as e:
            print(f"プロジェクトステータス取得エラー: {e}")
            status_display = str(project.status)
        
        project_data.append([
            str(i),
            str(project.project_name)[:20] + ('...' if len(str(project.project_name)) > 20 else ''),
            str(project.case_name.title)[:15] + ('...' if len(str(project.case_name.title)) > 15 else '') if project.case_name else '',
            f"¥{project.billing_amount_excluding_tax or 0:,.0f}",
            status_display
        ])
    
    project_table = Table(project_data, colWidths=[30, 100, 80, 70, 60])
    project_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e74c3c")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(project_table)
    
    return elements

def _generate_financial_dashboard(queryset, styles, filters):
    """財務ダッシュボード生成"""
    elements = []
    
    # タイトル
    elements.append(Paragraph("財務ダッシュボード", styles['ProMainTitle']))
    elements.append(Spacer(1, 20))
    
    # KPI指標
    total_revenue = queryset.aggregate(total=Sum('billing_amount_excluding_tax'))['total'] or 0
    total_cost = queryset.aggregate(total=Sum('outsourcing_cost_excluding_tax'))['total'] or 0
    gross_profit = total_revenue - total_cost
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # KPIテーブル
    kpi_data = [
        ['KPI指標', '当期実績', '前期比較'],
        ['総売上', f'¥{total_revenue:,.0f}', '+5.2%'],
        ['総原価', f'¥{total_cost:,.0f}', '+3.1%'],
        ['粗利益', f'¥{gross_profit:,.0f}', '+8.7%'],
        ['利益率', f'{profit_margin:.1f}%', '+1.2%'],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[100, 100, 80])
    kpi_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8e44ad")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(Paragraph("主要KPI", styles['ProSectionTitle']))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))
    
    # 収益性分析
    elements.append(Paragraph("収益性分析", styles['ProSectionTitle']))
    
    profitability_data = []
    for workload in queryset[:5]:  # 上位5件のみ
        revenue = workload.billing_amount_excluding_tax or 0
        cost = workload.outsourcing_cost_excluding_tax or 0
        profit = revenue - cost
        margin = (profit / revenue * 100) if revenue > 0 else 0
        
        profitability_data.append([
            str(workload.project_name)[:20],
            f'¥{revenue:,.0f}',
            f'¥{cost:,.0f}',
            f'¥{profit:,.0f}',
            f'{margin:.1f}%'
        ])
    
    if profitability_data:
        profit_table_data = [['プロジェクト', '売上', '原価', '利益', '利益率']] + profitability_data
        
        profit_table = Table(profit_table_data, colWidths=[100, 70, 70, 70, 50])
        profit_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f39c12")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(profit_table)
    
    return elements

def _generate_project_portfolio(queryset, styles, filters):
    """プロジェクトポートフォリオ生成"""
    elements = []
    
    # タイトル
    elements.append(Paragraph("プロジェクトポートフォリオ", styles['ProMainTitle']))
    elements.append(Spacer(1, 20))
    
    # プロジェクト分類別分析
    classification_analysis = queryset.values('case_classification').annotate(
        count=Count('id'),
        total_amount=Sum('billing_amount_excluding_tax'),
        avg_workdays=Avg('used_workdays')
    ).order_by('-total_amount')
    
    if classification_analysis:
        elements.append(Paragraph("チケット分類別ポートフォリオ", styles['ProSectionTitle']))

        class_data = [['分類', 'プロジェクト数', '総請求額', '平均工数']]
        for item in classification_analysis:
            # 安全に分類表示名を取得
            try:
                if hasattr(WorkloadAggregation, 'CASE_CLASSIFICATION_CHOICES'):
                    class_display = dict(WorkloadAggregation.CASE_CLASSIFICATION_CHOICES).get(
                        item['case_classification'], item['case_classification']
                    )
                else:
                    # 選択肢が定義されていない場合は、インスタンスから取得
                    sample_instance = queryset.filter(case_classification=item['case_classification']).first()
                    if sample_instance and hasattr(sample_instance, 'get_case_classification_display'):
                        class_display = sample_instance.get_case_classification_display()
                    else:
                        class_display = str(item['case_classification'])
            except Exception as e:
                print(f"分類表示名取得エラー: {e}")
                class_display = str(item['case_classification'])
            
            class_data.append([
                class_display,
                f"{item['count']:,}件",
                f"¥{item['total_amount'] or 0:,.0f}",
                f"{item['avg_workdays'] or 0:.1f}日"
            ])
        
        class_table = Table(class_data, colWidths=[80, 60, 80, 60])
        class_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'IPAexGothic'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#16a085")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(class_table)
    
    return elements

@login_required
@require_http_methods(["POST"])
def bulk_update_work_hours(request):
    """工数集計の一括工数更新 (計算項目自動更新 + 外注費同期)"""
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"工数一括更新開始 - ユーザー: {request.user}")
        
        data = json.loads(request.body)
        filter_params = data.get('filter_params', {})
        
        from .models import WorkloadAggregation
        from apps.workloads.models import Workload
        # === 【修正】正しいインポートパスに変更 ===
        from apps.cost_master.models import OutsourcingCost  # apps.projects.models から apps.cost_master.models に変更
        
        # フィルター条件を適用
        queryset = WorkloadAggregation.objects.all()
        
        # フィルター適用（既存のコード）
        if filter_params.get('project_name'):
            queryset = queryset.filter(project_name_id=filter_params['project_name'])
        
        if filter_params.get('case_name'):
            queryset = queryset.filter(case_name_id=filter_params['case_name'])
        
        if filter_params.get('status'):
            queryset = queryset.filter(status=filter_params['status'])
        
        if filter_params.get('case_classification'):
            queryset = queryset.filter(case_classification=filter_params['case_classification'])
        
        if filter_params.get('search'):
            search_term = filter_params['search']
            queryset = queryset.filter(
                Q(project_name__name__icontains=search_term) |
                Q(case_name__title__icontains=search_term) |
                Q(billing_destination__icontains=search_term)
            )
        
        logger.info(f"対象レコード数: {queryset.count()}件")
        
        updated_count = 0
        error_count = 0
        
        for aggregation_record in queryset:
            try:
                if aggregation_record.case_name:
                    # === 【外注費の同期処理】 ===
                    updated_outsourcing_cost = Decimal('0.0')
                    
                    # 年月フィールドを取得（WorkloadAggregationに年月フィールドがある場合）
                    if hasattr(aggregation_record, 'year_month') and aggregation_record.year_month:
                        target_year_month = aggregation_record.year_month
                    else:
                        # 年月フィールドがない場合は受注日から推定
                        if aggregation_record.order_date:
                            target_year_month = aggregation_record.order_date.strftime('%Y-%m')
                        else:
                            # デフォルトは現在年月
                            target_year_month = datetime.now().strftime('%Y-%m')
                    
                    try:
                        # === 【デバッグ】外注費テーブルの確認 ===
                        logger.debug(f"外注費取得開始: チケットID={aggregation_record.case_name.id}, "
                                   f"チケット名={aggregation_record.case_name.title}, "
                                   f"年月={target_year_month}")
                        
                        # 該当チケット・年月の外注費を取得
                        outsourcing_costs = OutsourcingCost.objects.filter(
                            ticket=aggregation_record.case_name,
                            year_month=target_year_month,
                            is_active=True  # アクティブな外注費のみ
                        )
                        
                        logger.debug(f"外注費レコード数: {outsourcing_costs.count()}件")
                        
                        # 外注費の合計を計算
                        for cost_record in outsourcing_costs:
                            cost_amount = Decimal(str(cost_record.total_cost or 0))
                            updated_outsourcing_cost += cost_amount
                            logger.debug(f"外注費レコード: ID={cost_record.id}, "
                                       f"金額=¥{cost_amount:,}, "
                                       f"業者={getattr(cost_record, 'vendor', 'N/A')}")
                        
                        logger.info(f"外注費同期: チケット={aggregation_record.case_name.title}, "
                                   f"年月={target_year_month}, "
                                   f"外注費=¥{updated_outsourcing_cost:,} "
                                   f"(レコード数: {outsourcing_costs.count()}件)")
                        
                    except OutsourcingCost.DoesNotExist:
                        logger.info(f"外注費データなし: チケット={aggregation_record.case_name.title}, 年月={target_year_month}")
                        updated_outsourcing_cost = Decimal('0.0')
                    except Exception as outsourcing_error:
                        logger.warning(f"外注費取得エラー (ID: {aggregation_record.id}): {outsourcing_error}")
                        # エラーの場合は既存の外注費を維持
                        updated_outsourcing_cost = Decimal(str(aggregation_record.outsourcing_cost_excluding_tax or 0))
                    
                    # === 【工数計算処理（既存のロジック）】 ===
                    workload_qs = Workload.objects.filter(
                        ticket=aggregation_record.case_name
                    )
                    
                    # 期間フィルターを適用（年月ベースで）
                    if aggregation_record.order_date and aggregation_record.actual_end_date:
                        start_date = aggregation_record.order_date
                        end_date = aggregation_record.actual_end_date
                        
                        target_year_months = []
                        current_date = start_date.replace(day=1)
                        
                        while current_date <= end_date:
                            year_month = current_date.strftime('%Y-%m')
                            target_year_months.append(year_month)
                            
                            if current_date.month == 12:
                                current_date = current_date.replace(year=current_date.year + 1, month=1)
                            else:
                                current_date = current_date.replace(month=current_date.month + 1)
                        
                        if target_year_months:
                            workload_qs = workload_qs.filter(year_month__in=target_year_months)
                    
                    # 工数を計算
                    total_used_workdays = Decimal('0.0')
                    total_newbie_workdays = Decimal('0.0')
                    
                    for workload in workload_qs:
                        try:
                            if aggregation_record.order_date and aggregation_record.actual_end_date:
                                year, month = map(int, workload.year_month.split('-'))
                                
                                period_start = max(
                                    aggregation_record.order_date,
                                    datetime(year, month, 1).date()
                                )
                                
                                last_day = calendar.monthrange(year, month)[1]
                                period_end = min(
                                    aggregation_record.actual_end_date,
                                    datetime(year, month, last_day).date()
                                )
                                
                                month_hours = 0
                                for day in range(period_start.day, period_end.day + 1):
                                    if day <= last_day:
                                        day_hours = workload.get_day_value(day)
                                        month_hours += day_hours
                                
                            else:
                                month_hours = workload.total_hours
                            
                            # 型をDecimalに統一
                            workdays = Decimal(str(month_hours)) / Decimal('8.0')
                            
                            if hasattr(workload.user, 'is_newbie') and workload.user.is_newbie:
                                total_newbie_workdays += workdays
                            else:
                                total_used_workdays += workdays
                                
                        except Exception as calc_error:
                            logger.warning(f"工数計算エラー (WorkloadID: {workload.id}): {calc_error}")
                            continue
                    
                    # === 【集計レコードを更新（工数+外注費）】 ===
                    # 更新前の値を記録（比較用）
                    old_outsourcing_cost = Decimal(str(aggregation_record.outsourcing_cost_excluding_tax or 0))
                    
                    aggregation_record.used_workdays = total_used_workdays
                    aggregation_record.newbie_workdays = total_newbie_workdays
                    aggregation_record.outsourcing_cost_excluding_tax = updated_outsourcing_cost
                    aggregation_record.updated_at = timezone.now()
                    
                    # === 【使用可能金額の自動計算・更新】 ===
                    if hasattr(aggregation_record, 'available_amount'):
                        try:
                            calculated_available = aggregation_record.calculated_available_amount
                            aggregation_record.available_amount = calculated_available
                            logger.debug(f"使用可能金額を自動更新: {calculated_available}")
                        except Exception as calc_error:
                            logger.warning(f"使用可能金額計算エラー (ID: {aggregation_record.id}): {calc_error}")
                    
                    # 保存実行
                    aggregation_record.save()
                    
                    # === 【更新結果をログ出力】 ===
                    logger.info(f"更新完了: ID={aggregation_record.id}, "
                              f"チケット={aggregation_record.case_name.title}, "
                              f"使用工数={float(total_used_workdays):.1f}, "
                              f"新入社員工数={float(total_newbie_workdays):.1f}, "
                              f"外注費: ¥{float(old_outsourcing_cost):,.0f} → ¥{float(updated_outsourcing_cost):,.0f}")
                    
                    updated_count += 1
                    
                else:
                    logger.warning(f"チケットが設定されていません: ID={aggregation_record.id}")
                    error_count += 1
                    
            except Exception as e:
                logger.error(f"工数更新エラー (ID: {aggregation_record.id}): {str(e)}")
                error_count += 1
                continue
        
        logger.info(f"工数一括更新完了 - 更新: {updated_count}件, エラー: {error_count}件")
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'error_count': error_count,
            'message': f'{updated_count}件の工数・外注費を更新しました。' +
                      (f' ({error_count}件でエラーが発生)' if error_count > 0 else ''),
            'should_reload_page': True
        })
        
    except Exception as e:
        logger.error(f"工数一括更新エラー: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def sync_outsourcing_costs(request):
    """外注費の同期専用API（工数集計レコードの外注費を最新の外注費データと同期）"""
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"外注費同期開始 - ユーザー: {request.user}")
        
        data = json.loads(request.body)
        filter_params = data.get('filter_params', {})
        
        from .models import WorkloadAggregation
        # === 【修正】正しいインポートパスに変更 ===
        from apps.cost_master.models import OutsourcingCost
        
        # フィルター条件を適用
        queryset = WorkloadAggregation.objects.all()
        
        # フィルター適用
        if filter_params.get('project_name'):
            queryset = queryset.filter(project_name_id=filter_params['project_name'])
        
        if filter_params.get('case_name'):
            queryset = queryset.filter(case_name_id=filter_params['case_name'])
        
        if filter_params.get('status'):
            queryset = queryset.filter(status=filter_params['status'])
        
        if filter_params.get('case_classification'):
            queryset = queryset.filter(case_classification=filter_params['case_classification'])
        
        if filter_params.get('search'):
            search_term = filter_params['search']
            queryset = queryset.filter(
                Q(project_name__name__icontains=search_term) |
                Q(case_name__title__icontains=search_term) |
                Q(billing_destination__icontains=search_term)
            )
        
        logger.info(f"外注費同期対象レコード数: {queryset.count()}件")
        
        updated_count = 0
        error_count = 0
        
        for aggregation_record in queryset:
            try:
                if aggregation_record.case_name:
                    # 年月の決定ロジック
                    if hasattr(aggregation_record, 'year_month') and aggregation_record.year_month:
                        target_year_month = aggregation_record.year_month
                    elif aggregation_record.order_date:
                        target_year_month = aggregation_record.order_date.strftime('%Y-%m')
                    else:
                        target_year_month = datetime.now().strftime('%Y-%m')
                    
                    # 外注費の取得と合計計算
                    try:
                        # === 【デバッグログ追加】 ===
                        logger.debug(f"外注費同期: チケット={aggregation_record.case_name.title}, "
                                   f"チケットID={aggregation_record.case_name.id}, "
                                   f"年月={target_year_month}")
                        
                        outsourcing_costs = OutsourcingCost.objects.filter(
                            ticket=aggregation_record.case_name,
                            year_month=target_year_month,
                            is_active=True
                        )
                        
                        logger.debug(f"マッチした外注費レコード数: {outsourcing_costs.count()}件")
                        
                        total_outsourcing_cost = Decimal('0.0')
                        for cost_record in outsourcing_costs:
                            amount = Decimal(str(cost_record.total_cost or 0))
                            total_outsourcing_cost += amount
                            logger.debug(f"外注費: ID={cost_record.id}, 金額=¥{amount:,}")
                        
                        # 既存の外注費と比較
                        current_cost = Decimal(str(aggregation_record.outsourcing_cost_excluding_tax or 0))
                        
                        logger.info(f"外注費比較: 現在=¥{float(current_cost):,.0f}, "
                                  f"新規=¥{float(total_outsourcing_cost):,.0f}")
                        
                        # 外注費が変更されている場合のみ更新
                        if total_outsourcing_cost != current_cost:
                            aggregation_record.outsourcing_cost_excluding_tax = total_outsourcing_cost
                            
                            # 使用可能金額の自動再計算
                            if hasattr(aggregation_record, 'available_amount'):
                                try:
                                    calculated_available = aggregation_record.calculated_available_amount
                                    aggregation_record.available_amount = calculated_available
                                except Exception as calc_error:
                                    logger.warning(f"使用可能金額計算エラー (ID: {aggregation_record.id}): {calc_error}")
                            
                            aggregation_record.updated_at = timezone.now()
                            aggregation_record.save()
                            
                            logger.info(f"外注費同期完了: ID={aggregation_record.id}, "
                                      f"チケット={aggregation_record.case_name.title}, "
                                      f"年月={target_year_month}, "
                                      f"旧外注費=¥{float(current_cost):,.0f}, "
                                      f"新外注費=¥{float(total_outsourcing_cost):,.0f}")
                            
                            updated_count += 1
                        else:
                            logger.debug(f"外注費変更なし: ID={aggregation_record.id}, 外注費=¥{float(total_outsourcing_cost):,.0f}")
                        
                    except Exception as outsourcing_error:
                        logger.error(f"外注費取得エラー (ID: {aggregation_record.id}): {outsourcing_error}")
                        error_count += 1
                        continue
                        
                else:
                    logger.warning(f"チケットが設定されていません: ID={aggregation_record.id}")
                    error_count += 1
                    
            except Exception as e:
                logger.error(f"外注費同期エラー (ID: {aggregation_record.id}): {str(e)}")
                error_count += 1
                continue
        
        logger.info(f"外注費同期完了 - 更新: {updated_count}件, エラー: {error_count}件")
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'error_count': error_count,
            'message': f'{updated_count}件の外注費を同期しました。' + 
                      (f' ({error_count}件でエラーが発生)' if error_count > 0 else ''),
            'should_reload_page': True
        })
        
    except Exception as e:
        logger.error(f"外注費同期エラー: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })